import openpyxl

path = r"C:\Users\U6033331\Desktop\Book1.xlsx"

workbook = openpyxl.load_workbook(path)

# sheet = workbook.get_sheet_by_name("Sheet1")
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column

for row in range(1, rows+1):
    for col in range(1, cols+1):
        print(sheet.cell(row=row,column=col).value, end = "     ")
    print()

for row in range(7,11):
    for col in range(1,4):
        sheet.cell(row=row, column=col).value = "Welcome" + str(row) + str(col)

workbook.save(path)